import re
import io
import json
import os
import time
import logging

from flask import Flask, render_template, request, send_file
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service

logging.basicConfig(level=logging.INFO)
app = Flask(__name__)

ONPE_URL = "https://consultaelectoral.onpe.gob.pe/inicio"
CHROME_BINARY   = os.getenv("CHROME_BINARY")
CHROMEDRIVER_PATH = os.getenv("CHROMEDRIVER_PATH")


def _build_driver():
    options = Options()
    options.add_argument("--headless=new")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1366,768")
    options.add_argument(
        "--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    )
    options.set_capability("goog:loggingPrefs", {"performance": "ALL"})

    if CHROME_BINARY:
        options.binary_location = CHROME_BINARY

    if CHROMEDRIVER_PATH:
        return webdriver.Chrome(service=Service(CHROMEDRIVER_PATH), options=options)

    # Selenium 4.6+ incluye selenium-manager que descarga el driver correcto automaticamente
    return webdriver.Chrome(options=options)


MEMBER_FIELDS = {
    "nombres", "cargo", "esMiembro", "region", "provincia", "distrito",
    "direccionLocal", "nombreCompleto", "local", "departamento",
    "ubigeoLocal", "localVotacion", "apellidoPaterno", "apellidoMaterno",
    "descripcionCargo", "descripcionMesa", "numeroMesa",
}


def _extract_from_logs(driver):
    """Busca en los performance logs la respuesta de la API de ONPE con datos del miembro."""
    logs = driver.get_log("performance")

    for entry in logs:
        try:
            msg = json.loads(entry["message"])["message"]
            if msg.get("method") != "Network.responseReceived":
                continue
            url = msg["params"]["response"]["url"]
            if "onpe.gob.pe" not in url:
                continue
            content_type = msg["params"]["response"].get("mimeType", "")
            if "json" not in content_type:
                continue
            req_id = msg["params"]["requestId"]
            body = driver.execute_cdp_cmd("Network.getResponseBody", {"requestId": req_id})
            resp = json.loads(body["body"])
            if not resp.get("success") or not resp.get("data"):
                continue
            d = resp["data"]
            if isinstance(d, list):
                if not d:
                    continue
                d = d[0]
            if not isinstance(d, dict):
                continue
            if MEMBER_FIELDS & set(d.keys()):
                return d
        except Exception:
            continue
    return None


def _parse_page_text(driver):
    """Fallback: extrae datos leyendo el texto visible de la página."""
    body_text = driver.find_element(By.TAG_NAME, "body").text
    lines = [l.strip() for l in body_text.splitlines() if l.strip()]
    upper = body_text.upper()

    if "NO ES MIEMBRO" in upper or "NO FIGURA" in upper or "NO SE ENCONTRO" in upper:
        return {"esMiembro": False}
    if "MIEMBRO" not in upper and "MESA" not in upper:
        return None

    def after(keyword):
        for i, line in enumerate(lines):
            if keyword.lower() in line.lower():
                if ":" in line:
                    val = line.split(":", 1)[1].strip()
                    if val:
                        return val
                if i + 1 < len(lines):
                    return lines[i + 1]
        return "-"

    return {
        "esMiembro":      True,
        "cargo":          after("cargo") if after("cargo") != "-" else after("rol"),
        "nombres":        after("nombre"),
        "region":         after("region") if after("region") != "-" else after("departamento"),
        "provincia":      after("provincia"),
        "distrito":       after("distrito"),
        "direccionLocal": after("local") if after("local") != "-" else after("direcci"),
    }


def consultar_onpe(dni):
    driver = None
    try:
        driver = _build_driver()
        driver.execute_cdp_cmd("Network.enable", {"maxResourceBufferSize": 10 * 1024 * 1024})
        driver.get(ONPE_URL)

        wait = WebDriverWait(driver, 25)
        dni_input = wait.until(EC.element_to_be_clickable(
            (By.CSS_SELECTOR, "input[type='text'], input[maxlength='8'], #numeroDni")
        ))
        dni_input.clear()
        dni_input.send_keys(dni)

        time.sleep(1)
        # Buscar el boton por texto primero, luego por fallback
        keywords_btn = ['consultar', 'buscar', 'verificar', 'search', 'submit', 'continuar']
        buttons = driver.find_elements(By.TAG_NAME, "button")
        submit = None
        for btn in buttons:
            if any(kw in btn.text.lower() for kw in keywords_btn):
                submit = btn
                break
        if not submit and buttons:
            submit = buttons[-1]
        if not submit:
            submit = driver.find_element(By.CSS_SELECTOR, "input[type='submit'], input[type='button']")
        driver.execute_script("arguments[0].click();", submit)

        time.sleep(6)

        data = _extract_from_logs(driver)
        if data:
            ubigeo = data.get("ubigeo") or ""
            parts = [p.strip() for p in ubigeo.split("/")]
            nombre_completo = f"{data.get('nombres', '')} {data.get('apellidos', '')}".strip()
            local_dir = " - ".join(filter(None, [
                data.get("localVotacion"), data.get("direccion")
            ])) or "-"
            return {
                "esMiembro":      data.get("miembroMesa", False),
                "cargo":          data.get("cargo", "-"),
                "nombres":        nombre_completo or "-",
                "region":         parts[0] if parts else "-",
                "provincia":      parts[1] if len(parts) > 1 else "-",
                "distrito":       parts[2] if len(parts) > 2 else "-",
                "direccionLocal": local_dir,
            }

        return _parse_page_text(driver)

    except Exception as e:
        err = str(e)
        logging.error(f"Error Selenium DNI {dni}: {err}")
        return {"error": err, "esMiembro": False}
    finally:
        if driver:
            driver.quit()


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/consultar', methods=['POST'])
def consultar():
    dnis_raw = request.form.get('dnis', '')
    dnis = [d.strip() for d in re.split(r'[\n,;\s]+', dnis_raw) if d.strip().isdigit()]
    resultados = []
    for dni in dnis:
        data = consultar_onpe(dni)
        if data and data.get('error'):
            resultados.append({
                'dni': dni,
                'es_miembro': False,
                'rol': 'ERROR',
                'nombres': data['error'][:120],
                'region': '-', 'provincia': '-', 'distrito': '-', 'direccion': '-',
            })
        elif data:
            resultados.append({
                'dni': dni,
                'es_miembro': data.get('esMiembro', False),
                'rol': data.get('cargo', '-'),
                'nombres': data.get('nombres', '-'),
                'region': data.get('region', '-'),
                'provincia': data.get('provincia', '-'),
                'distrito': data.get('distrito', '-'),
                'direccion': data.get('direccionLocal', '-'),
            })
        else:
            resultados.append({
                'dni': dni,
                'es_miembro': False,
                'rol': '-',
                'nombres': 'No se pudo consultar',
                'region': '-',
                'provincia': '-',
                'distrito': '-',
                'direccion': '-',
            })
    return render_template('index.html', resultados=resultados, dnis=dnis_raw)


@app.route('/descargar', methods=['POST'])
def descargar():
    data_json = request.form.get('data', '[]')
    resultados = json.loads(data_json)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Miembros de Mesa"

    headers = [
        'DNI', 'Rol (Miembro de Mesa)', 'Nombres',
        'Region', 'Provincia', 'Distrito', 'Direccion del Local'
    ]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for row, r in enumerate(resultados, 2):
        ws.cell(row=row, column=1, value=r['dni'])
        ws.cell(row=row, column=2, value=r['rol'])
        ws.cell(row=row, column=3, value=r['nombres'])
        ws.cell(row=row, column=4, value=r['region'])
        ws.cell(row=row, column=5, value=r['provincia'])
        ws.cell(row=row, column=6, value=r['distrito'])
        ws.cell(row=row, column=7, value=r['direccion'])

    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 4

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        download_name='miembros_mesa.xlsx',
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000)
